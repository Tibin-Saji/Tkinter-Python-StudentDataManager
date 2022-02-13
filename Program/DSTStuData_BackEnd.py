from openpyxl.workbook.workbook import Workbook
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Color, PatternFill
import os
from os import remove as osRemove
from math import isnan

DEBUG = False

if DEBUG:
    DSTFolderPath = 'DST\\'
    StudentFolderPath = os.path.join(os.path.dirname(os.path.dirname(__file__)), "Students\\")
    StudentListPath = 'Program\StudentsList.xlsx'
    Sub_TeachPath ='Program\\Subjects_Teachers.xlsx'
else:
    DSTFolderPath = 'DST\\'
    StudentFolderPath = 'DST\\Students\\'
    StudentListPath = DSTFolderPath + 'Program\\StudentList.xlsx'
    Sub_TeachPath = DSTFolderPath + 'Program\\Subjects_Teachers.xlsx'

ST = pd.DataFrame(pd.read_excel(Sub_TeachPath))

SUBJECTS = ST['Subject'].to_list()
TEACHERS = ST['Teacher'].to_list()

# StuList is the list of all the students added. It is a method to prevent future error like wrong name in marks or fees files
def UpdateStudentList():
    global StuList
    try:
        StuList = pd.DataFrame(pd.read_excel(StudentListPath))["ROLL NOS"].to_list()
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.cell(1,1).value = "ROLL NOS"
        ws.cell(1,2).value = "NAME OF THE STUDENT"
        wb.save(filename= StudentListPath)    # pathlib.Path(__file__).parent.resolve()
        StuList = pd.DataFrame(pd.read_excel(StudentListPath))["ROLL NOS"].to_list()


# Finds the row with empty first column. So should not have empty rows in between.
def max_row(ws):
    i = 0
    for rows in ws.iter_rows(values_only= True):
        i += 1
        if(rows[0] == None):        # If the first column of the row is empty 
            i -= 1      # This row is empty so does not belong to the table. So going back one row.
            break
    return i

# Name of the file cannot be with / but the roll no. is with /
def ConvertRollToText(roll):
    return roll.replace('/', '_')

# To remove any unwanted spaces in roll no.
def StringCorrection(name):
    name = name.replace(' ', '')
    return name

'''
*Parameter needed: path of the file.*
Converts the excel file in the path to a dataframe
'''
def GetFile(filePath):
    df = pd.read_excel(filePath)
    return df

'''
This function adds a row with the details about one subject. 
*Inputs needed : Marks file, Attendance file and the Subject name*
It will access each students excel file and append the data to the Marks worksheet.
'''
def AddMarks(pathM, pathA, subject, teacher):
    ReturnMsg = ''
    df = GetFile(pathM)

    t_Attendance = GetFile(pathA)

    new_header = t_Attendance.iloc[0]     # grab the first row for the header
    t_Attendance = t_Attendance[1:]     # take the data less the header row
    t_Attendance.columns = new_header     # set the header row as the df header
    t_Attendance.replace({'A':0,'P':1}, inplace = True)   # Give value to P and A
    t_Attendance = t_Attendance.sum(1)   # Add all the present days...It is a Series

    for index, rows in df.iterrows():
        rows[0] = StringCorrection(rows[0])
        
        # To check if the roll no. is empty or not
        if(not isinstance(rows[0], str)):
            continue
        # To check if the student roll exists in the student list 
        if(rows[0] not in StuList):
            ReturnMsg = ReturnMsg + rows[0] + ' does not exist in the Student List\n'
            continue

        studentFilePath = StudentFolderPath + ConvertRollToText(rows[0]) + '.xlsx'
        wb = openpyxl.load_workbook(filename= studentFilePath)
        ws = wb['Marks']

    #Adding data to Excel sheet
        noOfRows = max_row(ws)
        for i in range(noOfRows):
            if ws.cell(i + 1, 1).value == subject:
                noOfRows -= 1
                break

        
        
        ws.cell(noOfRows + 1, 1).value = subject                       # Subject
        ws.cell(noOfRows + 1, 1).alignment = Alignment(horizontal= 'center')
        ws.cell(noOfRows + 1, 2).value = teacher                       # Teacher
        ws.cell(noOfRows + 1, 2).alignment = Alignment(horizontal= 'center')
        ws.cell(noOfRows + 1, 3).value = rows[0]                       # Roll no
        ws.cell(noOfRows + 1, 3).alignment = Alignment(horizontal= 'center')
        ws.cell(noOfRows + 1, 4).value = t_Attendance.iloc[index]      # Attendance
        ws.cell(noOfRows + 1, 4).alignment = Alignment(horizontal= 'center')
        ws.cell(noOfRows + 1, 7).value = rows[4]                       # Attendance Mark
        ws.cell(noOfRows + 1, 7).alignment = Alignment(horizontal= 'center')
        ws.cell(noOfRows + 1, 8).value = rows[6]                       # Total Mark
        ws.cell(noOfRows + 1, 8).alignment = Alignment(horizontal= 'center')
        ws.cell(noOfRows + 1, 9).value = rows[5]                       # Grade
        ws.cell(noOfRows + 1, 9).alignment = Alignment(horizontal= 'center')
        
        fill=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('FF0000')) # Red Colour

        missedClass = False
        # Assignment Mark
        if isnan(float(rows[2])):
            missedClass = True
            ws.cell(noOfRows + 1, 5).value = rows[2] 
            ws.cell(noOfRows + 1, 5).fill = fill
        else:
            ws.cell(noOfRows + 1, 5).value = rows[2]
            ws.cell(noOfRows + 1, 5).alignment = Alignment(horizontal= 'center')
        
        # Exam Mark
        if isnan(float(rows[3])):
            missedClass = True
            ws.cell(noOfRows + 1, 6).value = rows[3]
            ws.cell(noOfRows + 1, 6).fill = fill
        else:
            ws.cell(noOfRows + 1, 6).value = rows[3]
            ws.cell(noOfRows + 1, 6).alignment = Alignment(horizontal= 'center')

            
        if(missedClass):
            ws.cell(noOfRows + 1, 9).fill = fill    # Grade
            ws.cell(noOfRows + 1, 8).fill = fill    # Total Marks

        wb.save(studentFilePath)

    ReturnMsg = ReturnMsg + 'Marks Added'
    return(ReturnMsg)


'''
This function adds a row with the details about the fees of one year. 
*Inputs needed : Fees file and the Year*
It will access each students excel file and append the data to the Marks worksheet.
'''
def AddFees(path, year):
    if(not(bool(path) and bool(year))):
        return ('Some fields are not entered\nNo change made')

    ReturnMsg = ''

    df = GetFile(path)

    noOfRows = 0

    for index, rows in df.iterrows():
        if (index == 0):
            continue
        # print(rows[0])
        # return
        rows[0] = str(rows[0])
        rows[0] = StringCorrection(rows[0])

        # To check if the roll no. is empty or not
        if(not isinstance(rows[0], str)):
            continue
        # To check if the student roll exists in the student list 
        if(rows[0] not in StuList):
            ReturnMsg = ReturnMsg + '%s does not exist in the Student List\n'%(rows[0])
            continue

        studentFilePath = StudentFolderPath + ConvertRollToText(rows[0]) + '.xlsx'
        wb = openpyxl.load_workbook(filename= studentFilePath)
        ws = wb['Fees']

        noOfRows = max_row(ws)   # Finds no. of rows to add the new line at the end

        flag = False
        rowToBeChanged = 0
        for i in range(max_row(ws)):
            if(ws.cell(i + 1, 1).value == year):     # cell(row based on 1, column based on 1)
                flag = True
                rowToBeChanged = i + 1
                break 
        
        
        if(not flag):
            rowToBeChanged = noOfRows + 1

        ws.cell(rowToBeChanged, 1).value = year
        ws.cell(rowToBeChanged, 1).alignment = Alignment(horizontal= 'center')
         
        rows[2] = 100 * rows[2]   #Converting the scholarship to percentage

        # Assigns and aligns for columns from B to F
        for i in range(2, 7):
            cell = ws.cell(rowToBeChanged, i)
            rows[i] = str(rows[i])
            if(rows[i] == 'nan'):
                rows[i] = 0
            if(i == 2):
                rows[2] = str(rows[2]) + '%' 
            cell.value = rows[i]
            cell.alignment = Alignment(horizontal= 'center')

            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

        wb.save(studentFilePath)

    ReturnMsg = ReturnMsg + 'Fees Added'
    return ReturnMsg

'''
*Inputs needed : the Subject to be deleted*
It will access each file and look for the subject name in each of them
and delete that line
It will print "Deleted" with the subject that was deleted
'''
def DeleteSubjectLine(subject):
    if(not bool(subject)):
        return ('No Subject selected\nNo change made')
    flag = False
    # ReturnMsg = ReturnMsg + rows[0] +
    # To access the worksheet named Subject in each excel file by checking the roll to the StuList
    for roll in StuList:
        studentFilePath = StudentFolderPath + ConvertRollToText(roll) + '.xlsx'

        wb = openpyxl.load_workbook(filename= studentFilePath)
        ws = wb['Marks']
        
        for i in range(max_row(ws)):
            if(ws.cell(i + 1, 1).value == subject):     # cell(row based on 1, column based on 1)
                flag = True
                ws.delete_rows(i + 1)       # delete_rows(row based on 1)
                break
        wb.save(studentFilePath)   

    if(flag):
            return('Marks of %s deleted'%(subject))
    else:
        return('No marks entered for %s'%(subject))

'''
*Inputs needed : the Year to be deleted*
It will access each file and look for the years in each of them
and delete that line
It will print "Deleted" with the year that was deleted
'''
def DeleteFeesLine(year):
    if(not bool(year)):
        return ('No Year entered\nNo change made')
    flag = False
    for roll in StuList:
        studentFilePath = StudentFolderPath + ConvertRollToText(roll) + '.xlsx'
        wb = openpyxl.load_workbook(filename= studentFilePath)
        ws = wb['Fees']
        for i in range(max_row(ws)):
            if(ws.cell(i + 1, 1).value == year):     # cell(row based on 1, column based on 1)
                flag = True
                ws.delete_rows(i + 1)       # delete_rows(row based on 1)
                break 

        wb.save(studentFilePath) 
    
    if(flag):
        return('Fees of %s deleted'%(year))
    else:
        return('No fees entered for %s'%(year))

def CreateMarksFilePerStudent(roll):
    
    # To check if the roll no. is empty or not
    if(not isinstance(roll, str)):
        return

    # List of the column headers
    of = [ 
                        'Subject',
                        'Teacher',
                        'Roll no.',
                        'Attendance',
                        'Assignment Mark',
                        'Exam Mark',
                        'Attendance Mark',
                        'Total Marks',
                        'Grade'
    ]

    studentFilePath = StudentFolderPath + ConvertRollToText(roll) + '.xlsx'

    # Open the already created file for student and create a sheet named Marks
    wb = openpyxl.load_workbook(studentFilePath)
    ws =  wb.create_sheet(title = 'Marks')
    
    ws.append(of)
    
    # To add each list from the list df to the ws and aligning each cell in 2nd column to center
    for i in range(len(of)):
        ws.cell(1, i + 1).alignment = Alignment(horizontal= 'center')
        ws.cell(1, i + 1).font = Font(b = True)

    # To change the width of the columns
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    wb.save(studentFilePath)

def CreateMarksFile():
    for roll in StuList:
        CreateMarksFilePerStudent(roll)

def CreateFeesFilePerStudent(roll):
    # To check if the roll no. is empty or not
    if(not isinstance(roll, str)):
        return
    # Dataframe of the column headers
    of = [
                    'Year',
                    'Scholarship',
                    'Fees to Pay (QR)',
                    'Fees to Pay (Rs)',
                    'Fees Paid (QR)',
                    'Fees Paid (Rs)',
    ]

    studentFilePath = StudentFolderPath + ConvertRollToText(roll) + '.xlsx'

    # Open the already created file for student and create a sheet named Fees
    wb = openpyxl.load_workbook(studentFilePath)
    ws =  wb.create_sheet(title = 'Fees')

    ws.append(of)

    # To add each list from the list df to the ws and aligning each cell in 2nd column to center
    for i in range(len(of)):
        ws.cell(1, i + 1).alignment = Alignment(horizontal= 'center')
        ws.cell(1, i + 1).font = Font(b = True)
    
    # To change the width of the columns
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16


    wb.save(studentFilePath)

def CreateFeesFile():
    for roll in StuList:
        CreateFeesFilePerStudent(roll)

def CreateMainDetailsFilePerStudent(of):
    roll = str(StringCorrection(of[1]))

    # List of list of rows to be added
    df = [
            ['Name: ', of[2]],
            ['Roll no.: ', roll],
            ['Contact no.: ', str(of[3])],
            ['Whatsapp no.: ', str(of[4])],
            ['Email Address: ', of[5]],
            ['Church: ', of[6]],
            ['Submited All Doc.: ', of[7]],
            ['Joining Year: ', str(of[1])[4:]],
            ['Graduation Year: ', '-']
    ]
    studentFilePath = StudentFolderPath  + ConvertRollToText(roll) + '.xlsx'   # Stores the path of a students file

    wb = openpyxl.Workbook()
    ws = wb.active      # Opens the first worksheet
    ws.title = "Main Details"       # Names it as Main Details

    
    # To add each list from the list df to the ws and aligning each cell in 2nd column to center
    for i in range(len(df)):
        ws.append(df[i])
        ws.cell(i + 1, 2).alignment = Alignment(horizontal= 'center')   # cell(row , column) row and column starts from 1. Alignment is a class
        ws.cell(i + 1, 1).font = Font(b= True)

    # To change the width of the first two columns
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30

    wb.save(studentFilePath)    # Save the workbook

def CreateMainDetailsFile():
    df = GetFile(input('Student Details'))     # Getting the Details file with details of all the students

    # Calling the func to create a file for each student
    for index, rows in df.iterrows():
        # To check if the roll no. is empty or not
        if(not isinstance(rows[1], str)):
            return
        # To check if the student roll exists in the student list 
        if(rows[1] not in StuList):
            print('\n%s does not exist in the Student List\n' %(rows[1]))
            return
        
        CreateMainDetailsFilePerStudent(rows)

'''
*Inputs needed: The updated excel file with all the details of the student*
Each student roll no in the details file is checked in the student list and if
it is not there, that student is given an excel file with filled out Details sheet
and empty Fees and Marks sheets.
It prints "Added" with the name of student and roll no. when added
'''
def AddStudents(path):
    ReturnMsg = ''
    df = GetFile(path)

    wb = openpyxl.load_workbook(filename= StudentListPath)
    ws = wb.active

    NewStu = 0

    for index, rows in df.iterrows():
        if(rows[1] not in StuList):
            NewStu = NewStu + 1
            # Creating complete Main Detail and blank Fee and Mark Files
            CreateMainDetailsFilePerStudent(rows)
            CreateMarksFilePerStudent(rows[1])
            CreateFeesFilePerStudent(rows[1])

            # Adding the student name and roll no. to the StudentList
            noOfRows = max_row(ws)

            # To add each list from the list df to the ws and aligning each cell in 2nd column to center
            ws.cell(noOfRows + 1, 1).value = rows[1]     # Roll no.
            ws.cell(noOfRows + 1, 1).alignment = Alignment(horizontal= 'center')   # cell(row , column) row and column starts from 1. Alignment is a class
            ws.cell(noOfRows + 1, 2).value = rows[2]     # Name
            ws.cell(noOfRows + 1, 2).alignment = Alignment(horizontal= 'center')

            ReturnMsg = ReturnMsg + "Added " + rows[2] + ' ' + rows[1] + '\n'

    wb.save(StudentListPath)

    UpdateStudentList()

    if(NewStu == 0):
        return 'No new Students to add'

    return ReturnMsg

'''
*Inputs needed: the roll no. of the student to be deleted*
Deletes the file of the student and also removes it from the student list
It prints "Deleted" with the roll no. of the student when deleted.
'''
def DeleteStudent(roll):
    if(roll not in StuList):
        return (roll + " does not exist \n")

    wb = openpyxl.load_workbook(StudentListPath)
    ws =  wb.active
    studentFilePath = StudentFolderPath + StringCorrection(ConvertRollToText(roll)) + '.xlsx'
    osRemove(studentFilePath)
    ws.delete_rows(StuList.index(roll) + 2, 1)
    StuList.pop(StuList.index(roll))
    wb.save(StudentListPath)

    UpdateStudentList()

    return (roll + " Deleted")

'''
This function should not be accessed by ppl. The initial 64 files should be made in 
our computer and the folder should be sent to them.
The CreateFunc if used will there is a file, will erase all data
'''
def CreateFunc():
    CreateMainDetailsFile()
    CreateMarksFile()
    CreateFeesFile()

UpdateStudentList()

